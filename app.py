import os
import json
import io
import csv
import uuid
from datetime import datetime
import markdown
from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
import anthropic
import openpyxl
from dv360_mapper import map_to_dv360, DV360_IO_COLUMNS

app = FastAPI()
templates = Jinja2Templates(directory="templates")
client = anthropic.Anthropic()

TTD_TEMPLATE_PATH = os.path.expanduser("~/Downloads/TTD BULKSHEET.xlsx")
DEFAULTS_PATH          = os.path.join(os.path.dirname(__file__), "defaults.json")
FEEDBACK_PATH          = os.path.join(os.path.dirname(__file__), "feedback.json")
MAPPING_PATH           = os.path.join(os.path.dirname(__file__), "MAPPING_REFERENCE.md")
PLATFORM_DEFAULTS_PATH = os.path.join(os.path.dirname(__file__), "platform_defaults.json")

# In-memory session store: session_id -> source_data
sessions = {}

TTD_SCHEMA = {
    "CampaignSets": [
        "IO ID", "Campaign Set ID", "Campaign Set Name"
    ],
    "Campaigns": [
        "Campaign Name", "Description", "Objective", "Primary Channel",
        "Goals", "Time Zone ID", "Pacing Mode", "IO Contract", "Campaign PO #"
    ],
    "Ad Groups": [
        "Ad Group Name", "Channel", "Goal Type", "Goal Value",
        "Base Bid", "Max Bid", "Marketplace", "Audience"
    ],
    "Budget Flights": [
        "Campaign", "Ad Group", "Flight Budget (in advertiser currency)",
        "Daily Spend Cap (in advertiser currency)", "Impression Budget",
        "Daily Impression Cap", "Start Date Inclusive UTC", "End Date Exclusive UTC", "Action"
    ],
    "Campaign Fees": ["Fee Name", "Value"],
    "Ad Group Fees": ["Ad Group Fee Name", "Value"]
}

SYSTEM_PROMPT = """You are an expert programmatic advertising specialist who translates media briefs into The Trade Desk (TTD) campaign bulk upload sheets.

You will receive data from 4 Excel input files:
1. Media Brief - campaign objectives, KPIs, targeting details, brand guidelines
2. Media Plan - channel breakdowns, flight dates, budgets, impressions, CPMs, DSP
3. Audience Matrix - audience segments, targeting types, segment descriptions, activation/suppression
4. Trafficking Sheet - campaign structure, creative details, flight information, city targeting

Map this data to the TTD bulk upload format and return valid JSON only."""


def load_defaults() -> dict:
    with open(DEFAULTS_PATH) as f:
        return json.load(f)


def load_feedback() -> list:
    if not os.path.exists(FEEDBACK_PATH):
        return []
    with open(FEEDBACK_PATH) as f:
        data = json.load(f)
    return data.get("rules", [])


def save_feedback_rule(rule: dict):
    if os.path.exists(FEEDBACK_PATH):
        with open(FEEDBACK_PATH) as f:
            data = json.load(f)
    else:
        data = {"rules": []}
    data["rules"].append(rule)
    with open(FEEDBACK_PATH, "w") as f:
        json.dump(data, f, indent=2)


def excel_to_dict(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(row)]
            else:
                if any(cell is not None for cell in row):
                    row_dict = {headers[j]: v for j, v in enumerate(row) if v is not None}
                    if row_dict:
                        rows.append(row_dict)
        result[sheet_name] = {"headers": headers, "rows": rows}
    return result


def create_ttd_excel(ttd_data: dict) -> bytes:
    with open(TTD_TEMPLATE_PATH, "rb") as f:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))

    sheet_key_map = {
        "campaign_sets": "CampaignSets",
        "campaigns": "Campaigns",
        "ad_groups": "Ad Groups",
        "budget_flights": "Budget Flights",
        "campaign_fees": "Campaign Fees",
        "ad_group_fees": "Ad Group Fees"
    }

    for data_key, sheet_name in sheet_key_map.items():
        rows = ttd_data.get(data_key, [])
        if not rows or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for i, row_data in enumerate(rows, start=2):
            for j, header in enumerate(headers, start=1):
                if header in row_data:
                    ws.cell(row=i, column=j, value=row_data[header])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def load_platform_defaults() -> dict:
    with open(PLATFORM_DEFAULTS_PATH) as f:
        return json.load(f)


def build_mapping_prompt(files_data: dict, extra_instruction: str = "") -> str:
    defaults          = load_defaults()
    platform_defaults = load_platform_defaults()
    feedback_rules    = load_feedback()

    feedback_section = ""
    if feedback_rules:
        feedback_section = f"""
LEARNED CORRECTIONS — apply these rules. They were confirmed correct by a human reviewer:
{json.dumps(feedback_rules, indent=2)}
"""

    revision_section = ""
    if extra_instruction:
        revision_section = f"""
REVISION REQUEST — the human reviewer flagged the following issue with the previous output. Fix it:
"{extra_instruction}"
"""

    return f"""Here is the data extracted from the 4 input files:

{json.dumps(files_data, indent=2, default=str)}

Map this to the TTD bulk upload format. Return a JSON object with exactly these keys:
- campaign_sets: list of row dicts for CampaignSets tab
- campaigns: list of row dicts for Campaigns tab
- ad_groups: list of row dicts for Ad Groups tab
- budget_flights: list of row dicts for Budget Flights tab
- campaign_fees: list of row dicts for Campaign Fees tab (empty list if none)
- ad_group_fees: list of row dicts for Ad Group Fees tab (empty list if none)

TTD field names to use (skip any marked [Read Only]):
{json.dumps(TTD_SCHEMA, indent=2)}

DEFAULT VALUES — apply in this priority order (most specific wins):
1. platform_defaults — TTD technical settings not in source documents (base layer)
2. global — applies to everything
3. by_channel — applies when channel is known
4. by_lob — applies when line of business is known
5. by_lob_and_channel — most specific, overrides all others

PLATFORM DEFAULTS (base layer — TTD technical fields):
{json.dumps(platform_defaults, indent=2)}

BUSINESS DEFAULTS (by LOB / channel):
{json.dumps(defaults, indent=2)}
{feedback_section}{revision_section}
Field mapping guidance:
- Skip all fields marked [Read Only]
- Dates (Start Date Inclusive UTC / End Date Exclusive UTC): format as "YYYY-MM-DD 00:00:00"
- Goal Type: map KPIs to TTD values (CPC, CPM, CPA, ROAS, VCR)
- Base Bid / Max Bid: dollar amounts in CPM
- Action (Budget Flights): "New" for new line items
- Each TTD line in the Media Plan becomes its own Ad Group and Budget Flight row
- Use audience segment names from the Audience Matrix in the Ad Group Audience field

Return ONLY valid JSON with no markdown, no explanation."""


def parse_claude_json(response_text: str) -> dict:
    if "```json" in response_text:
        response_text = response_text.split("```json")[1].split("```")[0].strip()
    elif "```" in response_text:
        response_text = response_text.split("```")[1].split("```")[0].strip()
    return json.loads(response_text)


def call_claude(prompt: str) -> dict:
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=16000,
        thinking={"type": "adaptive"},
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}]
    )
    response_text = next(b.text for b in response.content if b.type == "text")
    return parse_claude_json(response_text)


# ── Routes ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/mapping", response_class=HTMLResponse)
async def mapping_doc(request: Request):
    with open(MAPPING_PATH) as f:
        content = f.read()
    body = markdown.markdown(content, extensions=["tables"])
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>TTD Field Mapping Reference</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
           max-width: 1100px; margin: 0 auto; padding: 40px 24px;
           color: #1a1a2e; background: #f0f2f5; line-height: 1.6; }}
    .card {{ background: white; border-radius: 12px; padding: 40px;
             box-shadow: 0 2px 12px rgba(0,0,0,0.06); }}
    .nav {{ margin-bottom: 24px; font-size: 14px; }}
    .nav a {{ color: #1a73e8; text-decoration: none; }}
    .nav a:hover {{ text-decoration: underline; }}
    h1 {{ font-size: 24px; font-weight: 700; margin-bottom: 6px; }}
    h2 {{ font-size: 17px; font-weight: 700; margin: 36px 0 12px;
          padding-bottom: 8px; border-bottom: 2px solid #eee; color: #1a1a2e; }}
    h3 {{ font-size: 15px; font-weight: 600; margin: 24px 0 8px; }}
    p {{ margin: 0 0 12px; color: #444; font-size: 14px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-bottom: 24px; }}
    th {{ background: #f8f9fa; text-align: left; padding: 10px 12px;
          font-size: 11px; font-weight: 700; text-transform: uppercase;
          letter-spacing: 0.5px; color: #555; border-bottom: 2px solid #e0e0e0; }}
    td {{ padding: 9px 12px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }}
    tr:last-child td {{ border-bottom: none; }}
    tr:hover td {{ background: #f8f9ff; }}
    code {{ background: #f1f3f4; padding: 2px 6px; border-radius: 4px;
            font-size: 12px; font-family: monospace; }}
    blockquote {{ border-left: 3px solid #1a73e8; margin: 0 0 16px;
                  padding: 10px 16px; background: #e8f0fe; border-radius: 0 6px 6px 0; }}
    blockquote p {{ color: #1a1a2e; margin: 0; }}
    ul, ol {{ padding-left: 20px; margin: 0 0 12px; font-size: 14px; color: #444; }}
    li {{ margin-bottom: 4px; }}
    hr {{ border: none; border-top: 1px solid #eee; margin: 32px 0; }}
  </style>
</head>
<body>
  <div class="nav"><a href="/">← Back to app</a></div>
  <div class="card">{body}</div>
</body>
</html>""")



@app.get("/knowledge", response_class=HTMLResponse)
async def knowledge(request: Request):
    defaults          = load_defaults()
    platform_defaults = load_platform_defaults()
    feedback_rules    = load_feedback()
    return templates.TemplateResponse("knowledge.html", {
        "request":          request,
        "defaults":         defaults,
        "platform_defaults": platform_defaults,
        "feedback_rules":   feedback_rules,
    })


@app.post("/generate")
async def generate_ttd(
    media_brief: UploadFile = File(...),
    media_plan: UploadFile = File(...),
    audience_matrix: UploadFile = File(...),
    trafficking_sheet: UploadFile = File(...)
):
    files_data = {}
    for label, upload in [
        ("Media Brief", media_brief),
        ("Media Plan", media_plan),
        ("Audience Matrix", audience_matrix),
        ("Trafficking Sheet", trafficking_sheet),
    ]:
        content = await upload.read()
        files_data[label] = excel_to_dict(content)

    ttd_data = call_claude(build_mapping_prompt(files_data))

    session_id = str(uuid.uuid4())
    sessions[session_id] = files_data

    return JSONResponse({"session_id": session_id, "ttd_data": ttd_data})


@app.post("/revise")
async def revise_ttd(request: Request):
    body = await request.json()
    session_id = body["session_id"]
    revision_request = body["revision_request"]
    files_data = sessions.get(session_id, {})

    ttd_data = call_claude(build_mapping_prompt(files_data, extra_instruction=revision_request))

    # Ask Claude to extract a generalizable rule from this correction
    rule_response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": f"""A human reviewer corrected a TTD campaign mapping with this instruction:
"{revision_request}"

Extract a short, generalizable rule from this correction that can be applied to future campaigns.
Return a JSON object with these fields:
- rule: one-sentence rule (e.g. "CTV ad groups should always use VCR as Goal Type")
- field: the TTD field it applies to (e.g. "Goal Type")
- channel: channel it applies to, or "all" if universal
- lob: line of business it applies to, or "all" if universal

Return ONLY valid JSON, no markdown."""
        }]
    )

    try:
        rule_text = next(b.text for b in rule_response.content if b.type == "text")
        rule = parse_claude_json(rule_text)
        rule["date"] = datetime.now().strftime("%Y-%m-%d")
        rule["original_instruction"] = revision_request
        save_feedback_rule(rule)
    except Exception:
        pass  # Don't fail the revision if rule extraction fails

    return JSONResponse({"ttd_data": ttd_data})


@app.post("/export")
async def export_ttd(request: Request):
    body = await request.json()
    ttd_data = body["ttd_data"]
    excel_bytes = create_ttd_excel(ttd_data)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=TTD_Campaign_Bulk_Upload.xlsx"}
    )


# ── DV360 Routes ─────────────────────────────────────────────────────────────

@app.post("/generate/dv360")
async def generate_dv360(
    media_brief: UploadFile = File(...),
    media_plan: UploadFile = File(...),
    audience_matrix: UploadFile = File(...),
    trafficking_sheet: UploadFile = File(...)
):
    files_data = {}
    for label, upload in [
        ("Media Brief", media_brief),
        ("Media Plan", media_plan),
        ("Audience Matrix", audience_matrix),
        ("Trafficking Sheet", trafficking_sheet),
    ]:
        content = await upload.read()
        files_data[label] = excel_to_dict(content)

    dv360_data = map_to_dv360(files_data)

    session_id = str(uuid.uuid4())
    sessions[session_id] = files_data

    return JSONResponse({"session_id": session_id, "dv360_data": dv360_data})


def create_dv360_csv(dv360_data: dict) -> bytes:
    """Generate DV360 SDF v9.2 InsertionOrders CSV."""
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=DV360_IO_COLUMNS,
        extrasaction="ignore",
        lineterminator="\n"
    )
    writer.writeheader()
    for row in dv360_data.get("insertion_orders", []):
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


@app.post("/export/dv360")
async def export_dv360(request: Request):
    body = await request.json()
    dv360_data = body["dv360_data"]
    csv_bytes = create_dv360_csv(dv360_data)

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=InsertionOrders.csv"}
    )
