"""
Shared utilities imported by all three platform mappers (TTD, DV360, Amazon DSP).

Centralises every function that was previously copy-pasted across mapper.py,
dv360_mapper.py, and amazon_mapper.py.
"""

import io
import json
import os
from datetime import datetime
from typing import Optional

import openpyxl


# ── Channel normalisation ─────────────────────────────────────────────────────
# Keys must be lowercase — normalise_channel() lowercases before lookup.

CHANNEL_MAP = {
    "video ctv":             "CTV",
    "ctv":                   "CTV",
    "connected tv":          "CTV",
    "streaming tv":          "CTV",
    "connected home":        "Audio",
    "video olv":             "OLV",
    "olv":                   "OLV",
    "online video":          "OLV",
    "pre-roll":              "OLV",
    "display":               "Display",
    "banner":                "Display",
    "native":                "Native",
    "audio streaming audio": "Audio",
    "streaming audio":       "Audio",
    "audio podcasts":        "Audio",
    "audio":                 "Audio",
    "connected car":         "Audio",
    "dooh":                  "DOOH",
    "out of home":           "DOOH",
    "digital ooh":           "DOOH",
}


def normalise_channel(raw: str) -> str:
    """Map source channel strings to canonical channel names."""
    return CHANNEL_MAP.get(raw.strip().lower(), raw.strip())


# ── Input file parsers ────────────────────────────────────────────────────────

def parse_media_brief(sheet_data: dict) -> dict:
    """Media Brief: col A = label, col B = value. Returns flat dict."""
    rows = sheet_data.get("Sheet1", {}).get("rows", [])
    brief = {}
    for row in rows:
        values = list(row.values())
        if len(values) >= 2:
            label = str(values[0]).strip()
            value = values[1]
            if value is not None:
                brief[label] = value
    return brief


def parse_media_plan(sheet_data: dict, dsp_names: set) -> list:
    """
    Media Plan has two header rows; row 0 of rows[] is the real column header.
    Returns only rows where the DSP column matches dsp_names (case-insensitive).
    Pass an empty set to return all rows regardless of DSP.
    """
    raw_rows = sheet_data.get("Sheet1", {}).get("rows", [])
    if len(raw_rows) < 2:
        return []
    headers = list(raw_rows[0].values())
    lines = []
    for row in raw_rows[1:]:
        vals = list(row.values())
        record = dict(zip(headers, vals))
        dsp = str(record.get("DSP", "")).strip().lower()
        if not dsp_names or dsp in dsp_names:
            lines.append(record)
    return lines


def parse_trafficking_sheet(sheet_data: dict) -> list:
    """Returns rows from the Trafficking Sheet that have a Campaign value."""
    rows = sheet_data.get("Sheet1", {}).get("rows", [])
    return [r for r in rows if r.get("Campaign")]


# ── Name / LOB helpers ────────────────────────────────────────────────────────

def extract_lob(brief: dict, trafficking_row: dict = None) -> str:
    """Extract Line of Business from the Media Brief or Trafficking Sheet."""
    for key in ("LOB:", "LOB", "LOB/Corporate Function"):
        if brief.get(key):
            return str(brief[key]).strip()
    if trafficking_row:
        key = str(trafficking_row.get("Campaign Key", "")).strip()
        if key:
            return key
    return ""


def build_campaign_name(brief: dict, trafficking_rows: list,
                        fallback: str = "Unnamed Campaign") -> str:
    """
    Campaign name = first Campaign value in the Trafficking Sheet.
    Fallback: LOB + Product/Service from Media Brief.
    """
    for row in trafficking_rows:
        name = (row.get("Campaign") or "").strip()
        if name:
            return name
    parts = [
        brief.get("LOB:", brief.get("LOB", "")),
        brief.get("Product/Service:", brief.get("Product/Service", "")),
    ]
    return " - ".join(p for p in parts if p) or fallback


# ── Defaults helpers ──────────────────────────────────────────────────────────

def apply_platform_defaults(row: dict, tab: str, platform: dict) -> dict:
    """
    Merge platform defaults for a given tab into a row dict.
    Row values always win; blank platform default values ("") are skipped.
    """
    base = platform.get(tab, {})
    for field, value in base.items():
        if value != "" and field not in row:
            row[field] = value
    return row


def get_default(defaults: dict, field: str,
                channel: str = None, lob: str = None):
    """
    Most-specific-wins: by_lob_and_channel > by_channel > by_lob > global.
    """
    value = defaults.get("global", {}).get(field)

    if channel and channel in defaults.get("by_channel", {}):
        v = defaults["by_channel"][channel].get(field)
        if v is not None:
            value = v

    if lob and lob in defaults.get("by_lob", {}):
        v = defaults["by_lob"][lob].get(field)
        if v is not None:
            value = v

    if lob and channel:
        v = (defaults.get("by_lob_and_channel", {})
                     .get(lob, {}).get(channel, {}).get(field))
        if v is not None:
            value = v

    return value


# ── Date parsing ──────────────────────────────────────────────────────────────

# Source formats tried in order when parsing date strings.
_SRC_DATE_FMTS = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d")

# (start_format, end_format) per platform output format.
_OUTPUT_DATE_FMTS = {
    "ttd":    ("%Y-%m-%d 00:00:00", "%Y-%m-%d 00:00:00"),
    "dv360":  ("%m/%d/%Y",          "%m/%d/%Y"),
    "amazon": ("%m-%d-%Y-00-01",    "%m-%d-%Y-23-59"),
}


def parse_flight_dates(flight_str,
                       output_format: str = "ttd") -> tuple[Optional[str], Optional[str]]:
    """
    Parse flight date strings from the Media Plan 'Flight' column.
    Confirmed source format: 'M/D/YYYY - M/D/YYYY'  e.g. '12/1/2025 - 12/31/2025'
    Also handles datetime objects returned by openpyxl for date-formatted cells.

    output_format:
      "ttd"    → ("YYYY-MM-DD 00:00:00", "YYYY-MM-DD 00:00:00")
      "dv360"  → ("MM/DD/YYYY", "MM/DD/YYYY")
      "amazon" → ("MM-DD-YYYY-00-01", "MM-DD-YYYY-23-59")

    Returns (start_str, end_str); either may be None if not parseable.
    """
    if not flight_str:
        return None, None

    start_fmt, end_fmt = _OUTPUT_DATE_FMTS.get(output_format, _OUTPUT_DATE_FMTS["ttd"])

    if isinstance(flight_str, datetime):
        return flight_str.strftime(start_fmt), None

    s = str(flight_str).strip()

    def _to(p: str, fmt: str) -> Optional[str]:
        p = p.strip()
        for src in _SRC_DATE_FMTS:
            try:
                return datetime.strptime(p, src).strftime(fmt)
            except ValueError:
                continue
        return None

    if " - " in s:
        a, b = s.split(" - ", 1)
        return _to(a, start_fmt), _to(b, end_fmt)

    return _to(s, start_fmt), None


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def excel_to_dict(file_bytes: bytes) -> dict:
    """Convert Excel file bytes → {sheet_name: {headers: [...], rows: [...]}}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [
                    str(h).strip() if h is not None else f"col_{j}"
                    for j, h in enumerate(row)
                ]
            else:
                if any(cell is not None for cell in row):
                    row_dict = {headers[j]: v for j, v in enumerate(row) if v is not None}
                    if row_dict:
                        rows.append(row_dict)
        result[sheet_name] = {"headers": headers, "rows": rows}
    return result
