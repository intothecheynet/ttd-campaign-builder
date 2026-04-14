"""
TTD Input Filter.

Strips all non-TradeDesk content from the 4 input Excel files before
they are passed to the campaign builder. Produces two outputs per file:
  1. A cleaned Excel file — same structure as the original, TTD rows only.
  2. A filtered JSON dict — ready to pass directly into the mapper/pipeline.

How filtering works
───────────────────
Each sheet in each file is scanned for a "DSP column" — any column whose
header contains a DSP-related keyword (see DSP_COLUMN_KEYWORDS below).

If a DSP column is found:
  - Rows where that column contains a known TTD value are kept.
  - All other rows are removed.
  - The sheet's header rows are always preserved.

If no DSP column is found:
  - The sheet is kept in full (safe default — e.g. the Media Brief,
    which is campaign-level and applies to all DSPs).

DSP values recognised as TTD (case-insensitive):
  "ttd", "the trade desk", "tradedesk", "trade desk"
"""

import io
import json
from copy import copy
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter


# ── TTD recognition ───────────────────────────────────────────────────────────

TTD_VALUES = {"ttd", "the trade desk", "tradedesk", "trade desk"}

# Column header keywords that indicate a DSP/platform discriminator column.
# Matched case-insensitively against the header cell value.
DSP_COLUMN_KEYWORDS = {"dsp", "platform", "activation platform", "buying platform", "media platform"}


# ── Core filtering logic ──────────────────────────────────────────────────────

def _is_ttd_value(value) -> bool:
    """Return True if a cell value is a recognised TTD identifier."""
    if value is None:
        return False
    return str(value).strip().lower() in TTD_VALUES


def _find_dsp_column(headers: list) -> Optional[int]:
    """
    Return the 0-based index of the first column whose header matches a
    DSP column keyword. Returns None if no such column exists.
    """
    for i, h in enumerate(headers):
        if h and any(kw in str(h).strip().lower() for kw in DSP_COLUMN_KEYWORDS):
            return i
    return None


def _filter_sheet_rows(ws) -> tuple[list, int, int]:
    """
    Walk a worksheet and decide which rows to keep.

    Returns:
        kept_rows   — list of row tuples (values only) to write to output
        total_rows  — data rows examined (excludes header)
        removed_rows— data rows removed
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], 0, 0

    # Row 0 is always the header
    header = list(all_rows[0])
    dsp_col = _find_dsp_column(header)

    kept    = [header]   # always keep header
    total   = 0
    removed = 0

    for row in all_rows[1:]:
        if not any(cell is not None for cell in row):
            continue  # skip blank rows

        total += 1

        if dsp_col is not None:
            cell_val = row[dsp_col] if dsp_col < len(row) else None
            if _is_ttd_value(cell_val):
                kept.append(row)
            else:
                removed += 1
        else:
            # No DSP column — keep everything
            kept.append(row)

    return kept, total, removed


# ── Excel output ──────────────────────────────────────────────────────────────

def filter_to_excel(file_bytes: bytes, filename: str = "file") -> tuple[bytes, dict]:
    """
    Filter an Excel file to TTD-only rows.

    Returns:
        excel_bytes  — bytes of the cleaned Excel workbook
        summary      — dict describing what was filtered per sheet
    """
    wb_in  = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default empty sheet

    summary = {"file": filename, "sheets": []}

    for sheet_name in wb_in.sheetnames:
        ws_in = wb_in[sheet_name]
        kept_rows, total, removed = _filter_sheet_rows(ws_in)

        ws_out = wb_out.create_sheet(title=sheet_name)
        for row_data in kept_rows:
            ws_out.append(list(row_data))

        summary["sheets"].append({
            "sheet":        sheet_name,
            "rows_total":   total,
            "rows_removed": removed,
            "rows_kept":    total - removed,
            "filtered":     removed > 0,
        })

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output.read(), summary


# ── JSON output ───────────────────────────────────────────────────────────────

def filter_to_json(file_bytes: bytes) -> dict:
    """
    Filter an Excel file to TTD-only rows and return as structured JSON.

    Output format (matches excel_to_dict in app.py):
      { sheet_name: { "headers": [...], "rows": [{col: val, ...}, ...] } }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kept_rows, _, _ = _filter_sheet_rows(ws)

        if not kept_rows:
            result[sheet_name] = {"headers": [], "rows": []}
            continue

        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(kept_rows[0])]
        rows = []
        for row in kept_rows[1:]:
            row_dict = {headers[j]: v for j, v in enumerate(row) if v is not None}
            if row_dict:
                rows.append(row_dict)

        result[sheet_name] = {"headers": headers, "rows": rows}

    return result


# ── Batch: filter all 4 input files at once ──────────────────────────────────

def filter_all_inputs(
    media_brief_bytes:      bytes,
    media_plan_bytes:       bytes,
    audience_matrix_bytes:  bytes,
    trafficking_sheet_bytes: bytes,
) -> dict:
    """
    Filter all 4 campaign input files and return both cleaned Excel bytes
    and filtered JSON for each file.

    Returns:
      {
        "media_brief":       { "excel": bytes, "json": dict, "summary": dict },
        "media_plan":        { "excel": bytes, "json": dict, "summary": dict },
        "audience_matrix":   { "excel": bytes, "json": dict, "summary": dict },
        "trafficking_sheet": { "excel": bytes, "json": dict, "summary": dict },
      }
    """
    inputs = {
        "media_brief":       media_brief_bytes,
        "media_plan":        media_plan_bytes,
        "audience_matrix":   audience_matrix_bytes,
        "trafficking_sheet": trafficking_sheet_bytes,
    }

    results = {}
    for label, raw_bytes in inputs.items():
        excel_bytes, summary = filter_to_excel(raw_bytes, filename=label)
        json_data            = filter_to_json(raw_bytes)
        results[label] = {
            "excel":   excel_bytes,
            "json":    json_data,
            "summary": summary,
        }

    return results
